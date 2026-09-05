from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO
import json
import re
from typing import Any
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session
from ..models import ImportBatch, ImportRow, TrackerRecord, User, WorkLog
from .audit import audit

SOURCE_TABLES = {
    "DQ Task Tracker": ["Ticket ID", "Date Started", "Date Ended", "Tester", "DQ Status", "Zephyr Upload", "Comments"],
    "Daily Report - FT": ["TICKET ID", "TESTER", "TASK", "PRIORITY", "DATE", "work log (hrs)", "Passed - TC", "Passed - Steps", "Failed - TC", "Failed - Steps", "DAILY COMMENTS"],
    "Daily Report - BT": ["TICKET ID", "TESTER", "TASK", "PRIORITY", "DATE", "Passed - TC", "Passed - Steps", "Failed - TC", "Failed - Steps", "DAILY COMMENTS"],
}

@dataclass
class ParsedRow:
    sheet: str
    row_number: int
    payload: dict[str, Any]
    errors: list[str]


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def parse_number(value: Any) -> float | None:
    text = clean_text(value)
    if text is None or text == "-":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    number = parse_number(value)
    return int(number) if number is not None else None


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text or text in ("-", "0", "N/A", "na", "None"):
        return None
    # If Excel serial number as string or float
    try:
        num = float(text)
        if 20000 < num < 60000:
            from datetime import timedelta
            # Excel base date (accounting for 1900 leap year bug)
            base = date(1899, 12, 30)
            return base + timedelta(days=int(num))
    except ValueError:
        pass
    for fmt in (
        "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
        "%d-%m-%Y", "%d-%b-%y", "%d/%m/%y", "%m/%d/%y",
        "%b %d, %Y", "%d %b %Y", "%Y/%m/%d"
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def normalize_status(value: Any) -> str:
    text = clean_text(value) or "Pending"
    lowered = text.lower()
    if "complete" in lowered:
        return "Completed"
    if "block" in lowered or "hold" in lowered:
        return "Blocked"
    if "progress" in lowered:
        return "In progress"
    if "withdraw" in lowered:
        return "Withdrawn"
    return text


def find_header(ws, expected: list[str]) -> tuple[int, dict[str, int]] | None:
    expected_norm = {name.lower(): name for name in expected}
    for row in range(1, min(ws.max_row, 20) + 1):
        mapping: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = clean_text(ws.cell(row, col).value)
            if value and value.lower() in expected_norm:
                mapping[expected_norm[value.lower()]] = col
        if len(mapping) >= max(3, min(len(expected), 5)):
            return row, mapping
    return None


def find_grid_header(grid: list[list[Any]], expected: list[str]) -> tuple[int, dict[str, int]] | None:
    expected_norm = {name.lower(): name for name in expected}
    for row_idx, row in enumerate(grid[:20]):
        mapping: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            value = clean_text(cell)
            if value and value.lower() in expected_norm:
                mapping[expected_norm[value.lower()]] = col_idx
        if len(mapping) >= max(3, min(len(expected), 5)):
            return row_idx, mapping
    return None


def parse_grid_data(sheets_data: dict[str, list[list[Any]]]) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    for sheet, expected in SOURCE_TABLES.items():
        if sheet not in sheets_data:
            continue
        grid = sheets_data[sheet]
        if not grid:
            continue
        header = find_grid_header(grid, expected)
        if not header:
            rows.append(ParsedRow(sheet, 0, {}, ["Required header row was not found"]))
            continue
        header_row_idx, mapping = header
        for row_idx in range(header_row_idx + 1, len(grid)):
            row = grid[row_idx]
            raw = {name: (row[col_idx] if col_idx < len(row) else None) for name, col_idx in mapping.items()}
            # Ignore empty/template rows that have no ticket ID and no tester
            if sheet == "DQ Task Tracker":
                if not clean_text(raw.get("Ticket ID")) and not clean_text(raw.get("Tester")):
                    continue
            else:
                if not clean_text(raw.get("TICKET ID")) and not clean_text(raw.get("TESTER")):
                    continue
            if all(clean_text(value) is None for value in raw.values()):
                continue
            errors: list[str] = []
            if sheet == "DQ Task Tracker":
                payload = {
                    "ticket_id": clean_text(raw.get("Ticket ID")),
                    "date_started": parse_date(raw.get("Date Started")),
                    "date_ended": parse_date(raw.get("Date Ended")),
                    "tester_name_raw": clean_text(raw.get("Tester")),
                    "status": normalize_status(raw.get("DQ Status")),
                    "zephyr_upload": clean_text(raw.get("Zephyr Upload")),
                    "comments": clean_text(raw.get("Comments")),
                }
                for field in ("ticket_id", "date_started", "tester_name_raw", "status"):
                    if not payload.get(field):
                        errors.append(f"Missing required field: {field}")
            else:
                payload = {
                    "ticket_id_raw": clean_text(raw.get("TICKET ID")),
                    "workstream": "BT" if sheet.endswith("BT") else "FT",
                    "tester_name_raw": clean_text(raw.get("TESTER")),
                    "task": clean_text(raw.get("TASK")),
                    "priority": (clean_text(raw.get("PRIORITY")) or "").title() or None,
                    "work_date": parse_date(raw.get("DATE")),
                    "work_log_hours": parse_number(raw.get("work log (hrs)")),
                    "passed_tc": parse_int(raw.get("Passed - TC")),
                    "passed_steps": parse_int(raw.get("Passed - Steps")),
                    "failed_tc": parse_int(raw.get("Failed - TC")),
                    "failed_steps": parse_int(raw.get("Failed - Steps")),
                    "daily_comments": clean_text(raw.get("DAILY COMMENTS")),
                }
                for field in ("ticket_id_raw", "tester_name_raw", "work_date"):
                    if not payload.get(field):
                        errors.append(f"Missing required field: {field}")
            rows.append(ParsedRow(sheet, row_idx + 1, payload, errors))
    return rows


def parse_workbook(content: bytes) -> list[ParsedRow]:
    wb = load_workbook(BytesIO(content), data_only=True, keep_vba=False)
    rows: list[ParsedRow] = []
    for sheet, expected in SOURCE_TABLES.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header = find_header(ws, expected)
        if not header:
            rows.append(ParsedRow(sheet, 0, {}, ["Required header row was not found"]))
            continue
        header_row, mapping = header
        for row_num in range(header_row + 1, ws.max_row + 1):
            raw = {name: ws.cell(row_num, col).value for name, col in mapping.items()}
            if all(clean_text(value) is None for value in raw.values()):
                continue
            errors: list[str] = []
            if sheet == "DQ Task Tracker":
                payload = {
                    "ticket_id": clean_text(raw.get("Ticket ID")),
                    "date_started": parse_date(raw.get("Date Started")),
                    "date_ended": parse_date(raw.get("Date Ended")),
                    "tester_name_raw": clean_text(raw.get("Tester")),
                    "status": normalize_status(raw.get("DQ Status")),
                    "zephyr_upload": clean_text(raw.get("Zephyr Upload")),
                    "comments": clean_text(raw.get("Comments")),
                }
                for field in ("ticket_id", "date_started", "tester_name_raw", "status"):
                    if not payload.get(field):
                        errors.append(f"Missing required field: {field}")
            else:
                payload = {
                    "ticket_id_raw": clean_text(raw.get("TICKET ID")),
                    "workstream": "BT" if sheet.endswith("BT") else "FT",
                    "tester_name_raw": clean_text(raw.get("TESTER")),
                    "task": clean_text(raw.get("TASK")),
                    "priority": (clean_text(raw.get("PRIORITY")) or "").title() or None,
                    "work_date": parse_date(raw.get("DATE")),
                    "work_log_hours": parse_number(raw.get("work log (hrs)")),
                    "passed_tc": parse_int(raw.get("Passed - TC")),
                    "passed_steps": parse_int(raw.get("Passed - Steps")),
                    "failed_tc": parse_int(raw.get("Failed - TC")),
                    "failed_steps": parse_int(raw.get("Failed - Steps")),
                    "daily_comments": clean_text(raw.get("DAILY COMMENTS")),
                }
                for field in ("ticket_id_raw", "tester_name_raw", "work_date"):
                    if not payload.get(field):
                        errors.append(f"Missing required field: {field}")
            rows.append(ParsedRow(sheet, row_num, payload, errors))
    return rows


def preview_import(content: bytes) -> dict[str, Any]:
    parsed = parse_workbook(content)
    return {
        "total_rows": len(parsed),
        "valid_rows": sum(1 for row in parsed if not row.errors),
        "invalid_rows": sum(1 for row in parsed if row.errors),
        "errors": [{"sheet": row.sheet, "row": row.row_number, "errors": row.errors} for row in parsed if row.errors][:200],
        "sheets": sorted({row.sheet for row in parsed}),
    }


def import_parsed_rows(db: Session, actor: User, file_name: str, file_hash: str, parsed: list[ParsedRow], mode: str) -> ImportBatch:
    if mode not in {"merge", "replace", "add"}:
        raise ValueError("Import mode must be merge, replace, or add")
    batch = ImportBatch(file_name=file_name, file_hash=file_hash, mode=mode, status="running", imported_by_id=actor.id)
    db.add(batch)
    db.flush()
    if mode == "replace":
        for record in db.execute(select(TrackerRecord).where(TrackerRecord.deleted_at.is_(None))).scalars():
            record.deleted_at = datetime.utcnow()
        for log in db.execute(select(WorkLog).where(WorkLog.deleted_at.is_(None))).scalars():
            log.deleted_at = datetime.utcnow()
    success = rejected = 0
    for parsed_row in parsed:
        row_status = "accepted" if not parsed_row.errors else "rejected"
        if parsed_row.errors:
            rejected += 1
        else:
            try:
                if parsed_row.sheet == "DQ Task Tracker":
                    payload = parsed_row.payload
                    existing = db.execute(select(TrackerRecord).where(TrackerRecord.ticket_id == payload["ticket_id"], TrackerRecord.deleted_at.is_(None))).scalar_one_or_none()
                    if mode == "add" and existing:
                        raise ValueError("Duplicate tracker ticket")
                    record = existing if existing and mode == "merge" else TrackerRecord(ticket_id=payload["ticket_id"])
                    record.date_started = payload["date_started"]
                    record.date_ended = payload["date_ended"]
                    record.tester_name_raw = payload["tester_name_raw"]
                    record.status = payload["status"]
                    record.zephyr_upload = payload["zephyr_upload"]
                    record.comments = payload["comments"]
                    record.source_sheet = parsed_row.sheet
                    record.source_row = parsed_row.row_number
                    record.updated_by_id = actor.id
                    record.version = (record.version or 0) + 1
                    if not existing or mode != "merge":
                        record.created_by_id = actor.id
                        db.add(record)
                else:
                    payload = parsed_row.payload
                    match = db.execute(select(TrackerRecord).where(TrackerRecord.ticket_id == payload["ticket_id_raw"], TrackerRecord.deleted_at.is_(None))).scalar_one_or_none()
                    db.add(WorkLog(tracker_record_id=match.id if match else None, source_sheet=parsed_row.sheet, source_row=parsed_row.row_number, created_by_id=actor.id, updated_by_id=actor.id, **payload))
                success += 1
            except Exception as exc:
                row_status = "rejected"
                parsed_row.errors.append(str(exc))
                rejected += 1
        db.add(ImportRow(import_id=batch.id, sheet_name=parsed_row.sheet, row_number=parsed_row.row_number, status=row_status, error_message="; ".join(parsed_row.errors) or None, raw_json=json.dumps(parsed_row.payload, default=str)))
    batch.successful_rows = success
    batch.rejected_rows = rejected
    batch.status = "completed_with_errors" if rejected else "completed"
    batch.completed_at = datetime.utcnow()
    audit(db, actor, "import", "import", batch.id, None, {"file_name": file_name, "mode": mode, "success": success, "rejected": rejected})
    db.commit()
    return batch


def import_workbook(db: Session, actor: User, file_name: str, content: bytes, mode: str) -> ImportBatch:
    parsed = parse_workbook(content)
    file_hash = sha256(content).hexdigest()
    return import_parsed_rows(db, actor, file_name, file_hash, parsed, mode)


def import_grid_data(db: Session, actor: User, source_name: str, sheets_data: dict[str, list[list[Any]]], mode: str) -> ImportBatch:
    parsed = parse_grid_data(sheets_data)
    json_bytes = json.dumps(sheets_data, sort_keys=True).encode("utf-8")
    file_hash = sha256(json_bytes).hexdigest()
    return import_parsed_rows(db, actor, source_name, file_hash, parsed, mode)
